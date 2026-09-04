import os
import requests
from flask import Flask, jsonify, request, render_template
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import re

app = Flask(__name__)

DRIVE_FILE_ID = "14_9A0gjPBokDpdbw4wYGTepTdQKjbe8u"
EXCEL_PATH = "planilha_atual.xlsx"

CACHE_CLIENTES = []

def extrair_numero(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    val_str = str(val).strip()
    if val_str.startswith('='): return 0.0 # Ignora fórmulas não calculadas
    
    val_str = re.sub(r'[^\d,\.-]', '', val_str)
    if '.' in val_str and ',' in val_str:
        val_str = val_str.replace('.', '').replace(',', '.')
    elif ',' in val_str:
        val_str = val_str.replace(',', '.')
    try:
        return float(val_str)
    except:
        return 0.0

def formatar_data(val):
    if not val: return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    return str(val).split(' ')[0] # Remove os 00:00:00 se houver

def carregar_cache():
    global CACHE_CLIENTES
    if not os.path.exists(EXCEL_PATH):
        baixar_planilha_do_drive()
    
    # data_only=True garante que ele tente pegar o valor final e não a fórmula em texto
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb.active
    
    CACHE_CLIENTES = []
    for row in range(5, sheet.max_row + 1):
        col_a = sheet.cell(row=row, column=1).value
        nome = sheet.cell(row=row, column=2).value
        
        if not nome or col_a is None or not str(col_a).strip().isdigit():
            continue
            
        CACHE_CLIENTES.append({
            "linha": row,
            "nome": str(nome).strip(),
            "contrato": str(sheet.cell(row=row, column=3).value or ""),
            "valor_pego": extrair_numero(sheet.cell(row=row, column=6).value),    # Coluna F
            "valor_parcela": extrair_numero(sheet.cell(row=row, column=8).value), # Coluna H
            "valor_pago": extrair_numero(sheet.cell(row=row, column=10).value),   # Coluna J
            "pendente": extrair_numero(sheet.cell(row=row, column=11).value),     # Coluna K
            "multas": extrair_numero(sheet.cell(row=row, column=12).value),       # Coluna L
            "data_inicio": formatar_data(sheet.cell(row=row, column=13).value),   # Coluna M
            "data_fim": formatar_data(sheet.cell(row=row, column=14).value)       # Coluna N
        })

def baixar_planilha_do_drive():
    url = f"https://drive.google.com/uc?export=download&id={DRIVE_FILE_ID}"
    session = requests.Session()
    response = session.get(url, stream=True)
    for key, value in response.cookies.items():
        if key.startswith('download_warning'):
            params = {'export': 'download', 'id': DRIVE_FILE_ID, 'confirm': value}
            response = session.get(url, params=params, stream=True)
            break
    if response.status_code == 200:
        with open(EXCEL_PATH, 'wb') as f:
            for chunk in response.iter_content(chunk_size=1024):
                if chunk:
                    f.write(chunk)
    carregar_cache()

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/buscar-cliente', methods=['GET'])
def buscar_cliente():
    nome_busca = request.args.get('nome', '').strip().lower()
    if not nome_busca:
        return jsonify({"erro": "Informe o nome do cliente"}), 400
    
    if not CACHE_CLIENTES:
        carregar_cache()
        
    for cliente in CACHE_CLIENTES:
        partes_nome = cliente['nome'].lower().split()
        if any(p.startswith(nome_busca) or nome_busca in p for p in partes_nome):
            return jsonify(cliente)
            
    return jsonify({"erro": "Cliente não encontrado!"}), 404

@app.route('/api/lancar-noite', methods=['POST'])
def lancar_noite():
    dados = request.json
    try:
        if not os.path.exists(EXCEL_PATH):
            baixar_planilha_do_drive()
            
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb.active
        
        linha = dados.get('linha')
        tipo = dados.get('tipo') 
        valor_num = extrair_numero(dados.get('valor', 0))
        texto = dados.get('texto', str(valor_num))
        
        CORES = {
            "PIX": PatternFill(start_color="6AA84F", end_color="6AA84F", fill_type="solid"),
            "ESPECIE": PatternFill(start_color="8E7CC3", end_color="8E7CC3", fill_type="solid"),
            "ATRASO": PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid"),
            "SEMANAL": PatternFill(start_color="1155CC", end_color="1155CC", fill_type="solid"),
            "NEGOCIADO": PatternFill(start_color="EA9999", end_color="EA9999", fill_type="solid")
        }
        
        col_out_vlr = sheet.max_column
        col_aberto = col_out_vlr - 1
        
        if tipo in ["PIX", "ESPECIE", "SEMANAL"]:
            # Se for semanal e tiver texto, lança o texto (ex: 75 - Sexta). Se não, lança o número.
            valor_final_planilha = texto if tipo == "SEMANAL" else valor_num
            sheet.cell(row=linha, column=col_out_vlr, value=valor_final_planilha)
            
            fill = CORES.get(tipo, CORES["PIX"])
            sheet.cell(row=linha, column=col_aberto).fill = fill
            sheet.cell(row=linha, column=col_out_vlr).fill = fill
            
            # Tenta atualizar fisicamente as células 10 e 11 se NÃO forem fórmulas
            pago_cell = sheet.cell(row=linha, column=10)
            if not str(pago_cell.value).startswith('='):
                pago_cell.value = extrair_numero(pago_cell.value) + valor_num
                
            pendente_cell = sheet.cell(row=linha, column=11)
            if not str(pendente_cell.value).startswith('='):
                pendente_cell.value = extrair_numero(pendente_cell.value) - valor_num
                if pendente_cell.value <= 0:
                    sheet.cell(row=linha, column=4, value=datetime.now().strftime("%d/%m/%Y"))

        elif tipo == "ATRASO":
            status = dados.get('status', 'A')
            sheet.cell(row=linha, column=col_out_vlr, value=status)
            fill = CORES["ATRASO"]
            sheet.cell(row=linha, column=col_aberto).fill = fill
            sheet.cell(row=linha, column=col_out_vlr).fill = fill

        elif tipo == "NEGOCIADO":
            sheet.cell(row=linha, column=col_out_vlr, value=texto)
            fill = CORES["NEGOCIADO"]
            sheet.cell(row=linha, column=col_aberto).fill = fill
            sheet.cell(row=linha, column=col_out_vlr).fill = fill

        wb.save(EXCEL_PATH)

        # ATUALIZA A MEMÓRIA NA HORA PARA O CELULAR SOMAR CORRETAMENTE!
        for cliente in CACHE_CLIENTES:
            if cliente['linha'] == linha:
                if tipo in ["PIX", "ESPECIE", "SEMANAL"]:
                    cliente['valor_pago'] += valor_num
                    cliente['pendente'] -= valor_num
                break
                
        return jsonify({"mensagem": "Lançamento realizado instantaneamente!"})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route('/api/lancar-manha', methods=['POST'])
def lancar_manha():
    dados = request.json
    try:
        baixar_planilha_do_drive()
        wb = openpyxl.load_workbook(EXCEL_PATH)
        
        aba_ativa = wb.active
        nova_data = dados.get('nova_data', datetime.now().strftime("%d/%m"))
        nova_aba = wb.copy_worksheet(aba_ativa)
        nova_aba.title = f"CAIXA 07 BSB {nova_data}"
        
        COR_QUITADO = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
        
        for row in range(5, nova_aba.max_row + 1):
            pendente = nova_aba.cell(row=row, column=11).value
            if pendente is not None and extrair_numero(pendente) <= 0:
                for col in range(2, 15):
                    nova_aba.cell(row=row, column=col).fill = COR_QUITADO
                    
        wb.save(EXCEL_PATH)
        carregar_cache()
        return jsonify({"mensagem": "Rotina da manhã executada com sucesso!"})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

try:
    carregar_cache()
    print("Planilha carregada na memória com sucesso!")
except Exception as e:
    print("Erro ao carregar cache inicial:", e)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))